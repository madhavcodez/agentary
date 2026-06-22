"""Export structured research data in CSV, JSON, and Excel formats."""

import csv
import io
import json
import logging
from datetime import datetime
from typing import Any
from uuid import UUID

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ...models.entity import Entity
from ...models.entity_collection import EntityCollection
from ...models.finding import Finding

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEADER_BORDER = Border(
    bottom=Side(style="thin", color="000000"),
    right=Side(style="thin", color="D0D0D0"),
)
_ALT_ROW_FILL = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

# CSV column definitions for findings
_FINDING_CSV_COLUMNS = [
    "title",
    "category",
    "content",
    "confidence",
    "source_type",
    "source_name",
    "source_url",
    "expert",
    "tags",
    "verified",
    "created_at",
]


class DataExporter:
    """Export structured research data in various formats."""

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    def export_findings_csv(
        self,
        mission_id: UUID,
        filters: dict | None,
        db: Session,
    ) -> bytes:
        """Export findings as a UTF-8 CSV (with BOM for Excel compatibility).

        Columns: title, category, content, confidence, source_type,
        source_name, source_url, expert, tags, verified, created_at.

        Filters (optional dict keys):
            category        - exact match
            confidence_min  - float, minimum confidence threshold
            source_type     - exact match
        """
        findings = self._query_findings(mission_id, filters, db)

        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=_FINDING_CSV_COLUMNS)
        writer.writeheader()

        for f in findings:
            writer.writerow(self._finding_to_csv_row(f))

        csv_bytes = buf.getvalue().encode("utf-8-sig")
        logger.info(
            "Exported %d findings as CSV for mission %s",
            len(findings),
            mission_id,
        )
        return csv_bytes

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    def export_findings_json(
        self,
        mission_id: UUID,
        filters: dict | None,
        db: Session,
    ) -> str:
        """Export findings as a pretty-printed JSON string."""
        findings = self._query_findings(mission_id, filters, db)

        payload = [self._finding_to_dict(f) for f in findings]
        result = json.dumps(payload, indent=2, default=str, ensure_ascii=False)
        logger.info(
            "Exported %d findings as JSON for mission %s",
            len(findings),
            mission_id,
        )
        return result

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------

    def export_findings_excel(
        self,
        mission_id: UUID,
        filters: dict | None,
        db: Session,
    ) -> bytes:
        """Export findings as an Excel workbook with multiple sheets.

        Sheets:
            - Summary       -- mission stats overview
            - All Findings  -- complete data table
            - By Category   -- one sheet per distinct category
            - Sources       -- deduplicated source list

        Features: auto-fit column widths, styled headers, alternating row
        colours, frozen header rows, auto-filters.
        """
        findings = self._query_findings(mission_id, filters, db)

        wb = openpyxl.Workbook()
        # Remove the default sheet created by openpyxl
        wb.remove(wb.active)

        self._build_summary_sheet(wb, mission_id, findings)
        self._build_all_findings_sheet(wb, findings)
        self._build_category_sheets(wb, findings)
        self._build_sources_sheet(wb, findings)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(
            "Exported %d findings as Excel for mission %s",
            len(findings),
            mission_id,
        )
        return output.read()

    # ------------------------------------------------------------------
    # Structured data export
    # ------------------------------------------------------------------

    def export_structured_data(
        self,
        mission_id: UUID,
        format: str,
        db: Session,
    ) -> bytes:
        """Export the structured_data JSONB payloads from findings.

        Args:
            format: one of ``"csv"``, ``"json"``, or ``"excel"``.

        Returns raw bytes suitable for an HTTP response body.
        """
        findings = (
            db.query(Finding)
            .filter(
                Finding.mission_id == mission_id,
                Finding.structured_data.isnot(None),
            )
            .order_by(Finding.created_at)
            .all()
        )

        records: list[dict[str, Any]] = []
        all_keys: list[str] = []
        seen_keys: set[str] = set()

        for f in findings:
            data = f.structured_data
            if not isinstance(data, dict):
                continue
            row = {"finding_title": f.title, "finding_category": f.finding_type}
            row.update(data)
            records.append(row)
            for key in data:
                if key not in seen_keys:
                    seen_keys.add(key)
                    all_keys.append(key)

        columns = ["finding_title", "finding_category", *all_keys]

        fmt = format.lower().strip()
        if fmt == "json":
            return json.dumps(
                records, indent=2, default=str, ensure_ascii=False
            ).encode("utf-8")

        if fmt == "csv":
            buf = io.StringIO()
            writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            for row in records:
                writer.writerow(
                    {k: self._flatten_value(row.get(k, "")) for k in columns}
                )
            return buf.getvalue().encode("utf-8-sig")

        if fmt == "excel":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Structured Data"

            # Header
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                self._style_header_cell(cell)

            # Data rows
            for row_idx, row in enumerate(records, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    val = self._flatten_value(row.get(col_name, ""))
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = _CELL_ALIGNMENT
                    if row_idx % 2 == 0:
                        cell.fill = _ALT_ROW_FILL

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            self._auto_column_width(ws)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()

        raise ValueError(
            f"Unsupported export format: {format!r}. Use 'csv', 'json', or 'excel'."
        )

    # ------------------------------------------------------------------
    # Entity collection CSV
    # ------------------------------------------------------------------

    def export_entity_collection_csv(
        self,
        collection_id: UUID,
        db: Session,
    ) -> bytes:
        """Export an entity collection as CSV.

        Each row represents one entity. Columns are the union of all keys
        found across every entity's ``canonical_data`` dict, plus standard
        entity fields (name, entity_type, description, aliases, tags).
        """
        collection = (
            db.query(EntityCollection)
            .filter(EntityCollection.id == collection_id)
            .first()
        )
        if collection is None:
            logger.warning("Entity collection %s not found", collection_id)
            return b""

        entity_ids = collection.entity_ids or []
        if not entity_ids:
            logger.info("Entity collection %s has no entities", collection_id)
            return b""

        entities = (
            db.query(Entity).filter(Entity.id.in_(entity_ids)).all()
        )

        # Discover all canonical_data keys across every entity
        canonical_keys: list[str] = []
        seen: set[str] = set()
        for entity in entities:
            if entity.canonical_data and isinstance(entity.canonical_data, dict):
                for key in entity.canonical_data:
                    if key not in seen:
                        seen.add(key)
                        canonical_keys.append(key)

        base_columns = ["name", "entity_type", "description", "aliases", "tags"]
        columns = [*base_columns, *canonical_keys]

        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()

        for entity in entities:
            row: dict[str, Any] = {
                "name": entity.name or "",
                "entity_type": entity.entity_type or "",
                "description": entity.description or "",
                "aliases": "; ".join(entity.aliases) if entity.aliases else "",
                "tags": "; ".join(entity.tags) if entity.tags else "",
            }
            if entity.canonical_data and isinstance(entity.canonical_data, dict):
                for key in canonical_keys:
                    row[key] = self._flatten_value(
                        entity.canonical_data.get(key, "")
                    )
            writer.writerow(row)

        logger.info(
            "Exported %d entities from collection %s as CSV",
            len(entities),
            collection_id,
        )
        return buf.getvalue().encode("utf-8-sig")

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _query_findings(
        self,
        mission_id: UUID,
        filters: dict | None,
        db: Session,
    ) -> list[Finding]:
        """Query findings with optional filters.

        Supported filter keys:
            category        - exact string match
            confidence_min  - minimum confidence (float, inclusive)
            source_type     - exact string match
        """
        query = db.query(Finding).filter(Finding.mission_id == mission_id)

        if filters:
            if filters.get("category"):
                query = query.filter(Finding.finding_type == filters["category"])
            if "confidence_min" in filters and filters["confidence_min"] is not None:
                try:
                    threshold = float(filters["confidence_min"])
                    query = query.filter(Finding.confidence >= threshold)
                except (TypeError, ValueError):
                    logger.warning(
                        "Invalid confidence_min filter value: %r",
                        filters["confidence_min"],
                    )
            if filters.get("source_type"):
                query = query.filter(Finding.source_type == filters["source_type"])

        return query.order_by(Finding.created_at).all()

    # -- Row / dict serialisation helpers --------------------------------

    @staticmethod
    def _finding_to_csv_row(f: Finding) -> dict[str, Any]:
        """Convert a Finding ORM object to a flat dict for CSV writing."""
        return {
            "title": f.title or "",
            "category": f.finding_type or "",
            "content": f.content or "",
            "confidence": f.confidence if f.confidence is not None else "",
            "source_type": f.source_type or "",
            "source_name": f.source_name or "",
            "source_url": f.source_url or "",
            "expert": str(f.expert_agent_id) if f.expert_agent_id else "",
            "tags": "; ".join(f.tags) if f.tags else "",
            "verified": "Yes" if f.verified else "No",
            "created_at": (
                f.created_at.isoformat() if f.created_at else ""
            ),
        }

    @staticmethod
    def _finding_to_dict(f: Finding) -> dict[str, Any]:
        """Convert a Finding ORM object to a JSON-serialisable dict."""
        return {
            "id": str(f.id),
            "mission_id": str(f.mission_id),
            "category": f.finding_type,
            "title": f.title,
            "content": f.content,
            "structured_data": f.structured_data,
            "source_type": f.source_type,
            "source_url": f.source_url,
            "source_name": f.source_name,
            "confidence": f.confidence,
            "verified": f.verified,
            "tags": f.tags or [],
            "expert_agent_id": str(f.expert_agent_id) if f.expert_agent_id else None,
            "created_at": (
                f.created_at.isoformat() if f.created_at else None
            ),
        }

    @staticmethod
    def _flatten_value(value: Any) -> str:
        """Flatten a value into a string suitable for a CSV/Excel cell."""
        if value is None:
            return ""
        if isinstance(value, (dict, list)):
            return json.dumps(value, default=str, ensure_ascii=False)
        return str(value)

    # -- Excel sheet builders -------------------------------------------

    def _build_summary_sheet(
        self,
        wb: openpyxl.Workbook,
        mission_id: UUID,
        findings: list[Finding],
    ) -> None:
        """Create the Summary sheet with mission stats."""
        ws = wb.create_sheet("Summary")

        categories: dict[str, int] = {}
        source_types: dict[str, int] = {}
        verified_count = 0
        total_confidence = 0.0

        for f in findings:
            cat = f.finding_type or "uncategorised"
            categories[cat] = categories.get(cat, 0) + 1
            st = f.source_type or "unknown"
            source_types[st] = source_types.get(st, 0) + 1
            if f.verified:
                verified_count += 1
            total_confidence += f.confidence if f.confidence is not None else 0.0

        avg_confidence = (
            total_confidence / len(findings) if findings else 0.0
        )

        summary_rows: list[tuple[str, Any]] = [
            ("Mission ID", str(mission_id)),
            ("Export Date", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
            ("Total Findings", len(findings)),
            ("Verified Findings", verified_count),
            ("Average Confidence", f"{avg_confidence:.2f}"),
            ("", ""),
            ("Category Breakdown", ""),
        ]
        for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
            summary_rows.append((f"  {cat}", count))

        summary_rows.append(("", ""))
        summary_rows.append(("Source Type Breakdown", ""))
        for st, count in sorted(source_types.items(), key=lambda x: -x[1]):
            summary_rows.append((f"  {st}", count))

        for row_idx, (label, value) in enumerate(summary_rows, start=1):
            label_cell = ws.cell(row=row_idx, column=1, value=label)
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            if label and not label.startswith("  "):
                label_cell.font = Font(bold=True)
            value_cell.alignment = Alignment(horizontal="left")

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 45

    def _build_all_findings_sheet(
        self,
        wb: openpyxl.Workbook,
        findings: list[Finding],
    ) -> None:
        """Create the All Findings data table sheet."""
        ws = wb.create_sheet("All Findings")
        headers = [
            "Title",
            "Category",
            "Content",
            "Confidence",
            "Source Type",
            "Source Name",
            "Source URL",
            "Tags",
            "Verified",
            "Created At",
        ]

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._style_header_cell(cell)

        # Write data rows
        for row_idx, f in enumerate(findings, start=2):
            values = [
                f.title or "",
                f.finding_type or "",
                f.content or "",
                f.confidence if f.confidence is not None else "",
                f.source_type or "",
                f.source_name or "",
                f.source_url or "",
                "; ".join(f.tags) if f.tags else "",
                "Yes" if f.verified else "No",
                f.created_at.strftime("%Y-%m-%d %H:%M") if f.created_at else "",
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = _CELL_ALIGNMENT
                if row_idx % 2 == 0:
                    cell.fill = _ALT_ROW_FILL

        ws.freeze_panes = "A2"
        if findings:
            ws.auto_filter.ref = ws.dimensions
        self._auto_column_width(ws)

    def _build_category_sheets(
        self,
        wb: openpyxl.Workbook,
        findings: list[Finding],
    ) -> None:
        """Create one sheet per category."""
        by_category: dict[str, list[Finding]] = {}
        for f in findings:
            cat = f.finding_type or "uncategorised"
            by_category.setdefault(cat, []).append(f)

        for cat, cat_findings in sorted(by_category.items()):
            # Sheet names are limited to 31 characters in Excel
            sheet_name = cat[:31]
            ws = wb.create_sheet(sheet_name)

            headers = [
                "Title",
                "Content",
                "Confidence",
                "Source",
                "Verified",
                "Created At",
            ]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                self._style_header_cell(cell)

            for row_idx, f in enumerate(cat_findings, start=2):
                values = [
                    f.title or "",
                    f.content or "",
                    f.confidence if f.confidence is not None else "",
                    f.source_name or f.source_url or "",
                    "Yes" if f.verified else "No",
                    f.created_at.strftime("%Y-%m-%d %H:%M") if f.created_at else "",
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = _CELL_ALIGNMENT
                    if row_idx % 2 == 0:
                        cell.fill = _ALT_ROW_FILL

            ws.freeze_panes = "A2"
            if cat_findings:
                ws.auto_filter.ref = ws.dimensions
            self._auto_column_width(ws)

    def _build_sources_sheet(
        self,
        wb: openpyxl.Workbook,
        findings: list[Finding],
    ) -> None:
        """Create a Sources sheet with deduplicated source entries."""
        ws = wb.create_sheet("Sources")
        headers = ["Source Name", "Source Type", "URL", "Findings Count"]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._style_header_cell(cell)

        # Deduplicate sources by (source_name, source_url) pair
        source_key_to_data: dict[tuple[str, str], dict[str, Any]] = {}
        for f in findings:
            name = f.source_name or ""
            url = f.source_url or ""
            if not name and not url:
                continue
            key = (name, url)
            if key in source_key_to_data:
                source_key_to_data[key]["count"] += 1
            else:
                source_key_to_data[key] = {
                    "name": name,
                    "type": f.source_type or "",
                    "url": url,
                    "count": 1,
                }

        sorted_sources = sorted(
            source_key_to_data.values(), key=lambda s: -s["count"]
        )

        for row_idx, source in enumerate(sorted_sources, start=2):
            values = [
                source["name"],
                source["type"],
                source["url"],
                source["count"],
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = _CELL_ALIGNMENT
                if row_idx % 2 == 0:
                    cell.fill = _ALT_ROW_FILL

        ws.freeze_panes = "A2"
        if sorted_sources:
            ws.auto_filter.ref = ws.dimensions
        self._auto_column_width(ws)

    # -- Styling helpers -------------------------------------------------

    @staticmethod
    def _style_header_cell(cell: openpyxl.cell.cell.Cell) -> None:
        """Apply standard header styling to an Excel cell."""
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _HEADER_BORDER

    @staticmethod
    def _auto_column_width(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        min_width: int = 10,
        max_width: int = 50,
    ) -> None:
        """Auto-fit column widths based on cell content."""
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    # Use the first line only for multi-line content
                    first_line = cell_value.split("\n")[0]
                    max_length = max(max_length, len(first_line))
                except (TypeError, AttributeError):
                    pass
            # Add a small buffer for padding
            adjusted = min(max(max_length + 3, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted
